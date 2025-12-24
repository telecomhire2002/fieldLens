# app/routes/jobs.py

from fastapi import APIRouter, Depends, HTTPException, Response, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from bson import ObjectId
from typing import List, Optional, Any, Dict
from datetime import datetime, date
import csv, io, os, zipfile, tempfile
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import httpx
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
import os
import re
from app.deps import get_db
from app.schemas import CreateJob, JobOut, PhotoOut
from app.models import new_job 
from app.services.storage_s3 import presign_url, get_bytes
from app.utils import normalize_phone, build_required_types_for_sector, type_label,sector_by_id

router = APIRouter()
TEMPLATE_BOOK1_PATH = os.path.join(os.path.dirname(__file__), "app/static/templates/Book1_template.xlsx"
)


def oid(obj):
    return str(obj["_id"]) if isinstance(obj.get("_id"), ObjectId) else obj.get("_id")


# ------------------------------------------------------------
# LIST JOBS
# ------------------------------------------------------------
def _job_to_out(doc: dict) -> JobOut:
    created = doc.get("createdAt")
    if isinstance(created, (datetime, date)):
        created = created.isoformat()

    # Normalize sectors into the schema shape
    sectors_out = []
    for s in (doc.get("sectors") or []):
        sectors_out.append(
            {
                "sector": str(s.get("sector")),
                "requiredTypes": s.get("requiredTypes", []),
                "currentIndex": int(s.get("currentIndex", 0)),
                "status": s.get("status", "PENDING"),
            }
        )

    return JobOut(
        id=str(doc["_id"]),
        workerPhone=doc["workerPhone"],
        siteId=doc["siteId"],
        sectors=sectors_out,
        requiredTypes=doc.get("requiredTypes", []),
        currentIndex=int(doc.get("currentIndex", 0) or 0),
        status=doc.get("status", "PENDING"),
        createdAt=created,
        macId=doc.get("macId"),
        rsnId=doc.get("rsnId"),
        azimuthDeg=doc.get("azimuthDeg"),
        circle=doc.get("circle"),
        company=doc.get("company"),
    )




@router.get("/jobs")
def list_jobs(db=Depends(get_db)) -> List[JobOut]:
    docs = list(db.jobs.find({}, sort=[("_id", -1)]))
    return [_job_to_out(d) for d in docs]


@router.get("/jobs/{job_id}")
def get_job(
    job_id: str,
    sector: Optional[int] = Query(None),
    db=Depends(get_db)
) -> Dict[str, Any]:
    """
    Return a single job + its photos.
    All photos include a presigned HTTPS URL (s3Url).
    """
    try:
        _id = ObjectId(job_id)
    except Exception:
        raise HTTPException(400, "Invalid job id")

    job = db.jobs.find_one({"_id": _id})
    if not job:
        raise HTTPException(404, "Job not found")

    photo_q: Dict[str, Any] = {"jobId": str(job["_id"])}
    if sector is not None:
        photo_q["sector"] = sector
    photos = list(db.photos.find(photo_q).sort("_id", 1))

    out_photos = []
    for p in photos:
        docp: Dict[str, Any] = {
            "id": str(p["_id"]),
            "jobId": p.get("jobId"),
            "type": p.get("type"),
            "sector": p.get("sector"),
            "status": p.get("status"),
            "reason": p.get("reason") or [],
            "fields": p.get("fields") or {},
            "checks": p.get("checks") or {},
            "phash": p.get("phash"),
            "ocrText": p.get("ocrText"),
            "s3Key": p.get("s3Key"),
        }

        key = p.get("s3Key")
        docp["s3Url"] = presign_url(key, expires=3600) if key else None

        out_photos.append(docp)

    return {
        "job": _job_to_out(job).model_dump(),
        "photos": out_photos,
    }


@router.post("/jobs", response_model=JobOut)
def create_or_extend_job(payload: CreateJob, db=Depends(get_db)) -> JobOut:
    """
    Create a job for a (workerPhone, siteId, sector) triple.

    We now store one Mongo document per SECTOR so that WhatsApp logic
    (which picks jobs by sector) works cleanly.

    The document contains:
    - top-level sector / requiredTypes / currentIndex (used by WhatsApp)
    - a 'sectors' array with a single SectorProgress block (used by UI)
    """
    worker = payload.workerPhone.strip()
    site = payload.siteId.strip()
    sector = payload.sector.strip()  # string

    if not worker or not site or not sector:
        raise HTTPException(400, "workerPhone, siteId and sector are required")

    # Store worker phone in the SAME canonical form WhatsApp uses
    worker_phone = normalize_phone(worker)

    # 1) If a job for this worker+site+sector already exists, just return it
    existing = db.jobs.find_one(
        {"workerPhone": worker_phone, "siteId": site, "sector": sector}
    )
    if existing:
        return _job_to_out(existing)

    # 2) Build the 14-step required types for this sector
    sector_required = build_required_types_for_sector(sector)

    # Sector block for UI
    sector_block = {
        "sector": sector,
        "requiredTypes": sector_required,
        "currentIndex": 0,
        "status": "PENDING",
    }

    # 3) Create job document – SINGLE sector per job
    doc: Dict[str, Any] = {
        "workerPhone": worker_phone,
        "siteId": site,
        "sector": sector,                  # <- used by WhatsApp
        "requiredTypes": sector_required,  # <- used by WhatsApp
        "currentIndex": 0,                 # <- used by WhatsApp

        "sectors": [sector_block],         # <- used by dashboard UI
        "status": "PENDING",
        "createdAt": datetime.utcnow().isoformat(),
        "circle": payload.circle,
        "company": payload.company,
    }

    ins = db.jobs.insert_one(doc)
    doc["_id"] = ins.inserted_id
    return _job_to_out(doc)



# ------------------------------------------------------------
# PER-JOB CSV
# ------------------------------------------------------------
@router.post("/jobs/{job_id}/export.csv")
def export_csv(
    job_id: str,
    mainExcel: UploadFile = File(...),
    db=Depends(get_db),
):
    # -------- 1. Fetch base job --------
    try:
        base_job = db.jobs.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(400, "Invalid Job ID")

    if not base_job:
        raise HTTPException(404, "Job not found")

    worker = base_job.get("workerPhone")
    site_id = str(base_job.get("siteId", "")).strip()
    circle = base_job.get("circle", "")

    # -------- 2. Collect all sectors (mac/rsn) for this site --------
    related = list(db.jobs.find({"workerPhone": worker, "siteId": site_id}))
    sec_info = []

    alpha_map = {"alpha": 1, "beta": 2, "gamma": 3}

    for j in related:
        raw_sec = str(j.get("sector", "")).strip()

        sec_num = None
        low = raw_sec.lower()
        if low in alpha_map:
            sec_num = alpha_map[low]
        elif raw_sec.startswith("-") and raw_sec[1:].isdigit():
            sec_num = int(raw_sec[1:])
        elif raw_sec.isdigit():
            sec_num = int(raw_sec)
        elif low.startswith("sec") and raw_sec[3:].isdigit():
            sec_num = int(raw_sec[3:])

        sector_norm = f"Sec{sec_num}" if sec_num else raw_sec
        azimuth = j.get("azimuthDeg") or ""
        mac = j.get("macId") or ""
        rsn = j.get("rsnId") or ""

        # fallback from photos
        for p in db.photos.find({"jobId": str(j["_id"])}):
            f = p.get("fields") or {}
            if not mac and f.get("macId"):
                mac = f["macId"]
            if not rsn and f.get("rsn"):
                rsn = f["rsn"]

        sec_info.append({"sector": sector_norm, "mac": mac, "rsn": rsn, "azimuth":azimuth})

    # -------- 3. Read uploaded Main Excel --------
    try:
        df = pd.read_excel(mainExcel.file)
    except Exception:
        raise HTTPException(400, "Failed to read uploaded Excel")

    cols = {c.lower(): c for c in df.columns}

    def find_c(*keys):
        for low, orig in cols.items():
            if all(k.lower() in low for k in keys):
                return orig
        return None

    site_col = find_c("enbsiteid") or find_c("site")
    pmp_col = find_c("pmp sap id", "sap")
    a6_col = find_c("a6neid")
    gis_col = find_c("gis sector_id", "sector")
    a6ip_col = find_c("a6 ip")  # IPv6 pool
    a6hieght_col = find_c("enb antenna height")
    a6tilt_col = find_c("proposed a6 tilt")
    sitename_col = find_c("site name")

    if not site_col:
        raise HTTPException(400, "Site / eNBsiteID not found in Main Excel")

    df_site = df[site_col].astype(str).str.strip()
    match = df[df_site == site_id]
    if match.empty:
        match = df[df_site.str.contains(site_id, case=False, na=False)]

    if match.empty:
        base_pmp = base_a6 = base_gis = base_a6ip = base_a6height = base_a6tilt = ""
    else:
        r = match.iloc[0]

        def safe(c):
            return "" if not c or c not in r or pd.isna(r[c]) else str(r[c])

        base_pmp = safe(pmp_col)
        base_a6 = safe(a6_col)
        base_gis = safe(gis_col)
        base_a6ip = safe(a6ip_col)
        base_a6height = safe(a6hieght_col)
        base_a6tilt = safe(a6tilt_col)
        base_sitename = safe(sitename_col)

    def _sec_sort_key(x):
        m = re.findall(r"\d+", str(x.get("sector", "")))
        return int(m[0]) if m else 999
    def as_int_str(v):
        """
        Converts 10 / 10.0 / '10.0' → '10'
        Returns '' for empty/NaN
        """
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
            return str(int(float(v)))
        except Exception:
            return str(v).strip()


    sec_info_sorted = sorted(sec_info, key=_sec_sort_key)
    sector_count = len(sec_info_sorted)

    # --- Azimuth: comma-separated across sectors (skip blanks) ---
    azimuth_values = [str(d.get("azimuth", "")).strip() for d in sec_info_sorted]
    azimuth_values = [v for v in azimuth_values if v]  # remove empty
    azimuth_combined = ", ".join(azimuth_values)

    # --- Height / Tilt: repeat same base value N times ---
    def repeat_base(val: str, n: int) -> str:
        v = as_int_str(val)
        if not v or n <= 0:
            return ""
        return ", ".join([v] * n)

    a6height_combined = repeat_base(base_a6height, sector_count)
    a6tilt_combined   = repeat_base(base_a6tilt, sector_count)



    # -------- 4. Helper – A6 & IPv6 per sector --------
    def a6_for_sector(sec: str | None) -> str:
        if not base_a6 or not sec:
            return ""

        m = re.findall(r"\d+", sec)
        if not m:
            return ""
        target_num = int(m[0])

        # all sector numbers present
        sector_nums = sorted(
            int(re.findall(r"\d+", d["sector"])[0]) for d in sec_info
            if re.findall(r"\d+", d["sector"])
        )
        count = len(sector_nums)

        t = re.search(r"(\d+)$", base_a6)
        if not t:
            return base_a6

        full_suffix = t.group(1)
        suffix_digit = int(full_suffix[-1])
        prefix = base_a6[:-len(full_suffix)]

        if count == 1:
            only = sector_nums[0]
            return base_a6 if only == target_num else ""

        if count == 2:
            s1, s2 = sector_nums
            if suffix_digit in sector_nums:
                if target_num == suffix_digit:
                    return base_a6
                return f"{prefix}{full_suffix[:-1]}{target_num}"
            return f"{prefix}{full_suffix[:-1]}{target_num}"

        # 3 sectors
        return f"{prefix}{full_suffix[:-1]}{target_num}"

    def a6ip_for_sector(sec: str | None) -> str:
        if not base_a6ip or not sec:
            return ""
        m = re.search(r"(\d+)$", base_a6ip)
        if not m:
            return base_a6ip
        base_last = int(m.group(1))

        sec_num = int(re.findall(r"\d+", sec)[0])
        sector_numbers = sorted(
            int(re.findall(r"\d+", d["sector"])[0])
            for d in sec_info
        )

        if len(sector_numbers) == 1:
            if sector_numbers[0] == sec_num:
                return base_a6ip
            return ""

        base_sec = sector_numbers[0]
        diff = sec_num - base_sec
        new_last = base_last + diff
        prefix = base_a6ip[:-len(str(base_last))]
        return prefix + str(new_last)

    # helper: sector from Column A text
    def sector_from_hc(hc: str):
        if not hc:
            return None
        m = re.search(r"sect\s*([0-9]+)", hc.lower())
        if m:
            return f"Sec{m.group(1)}"
        return None


    # -------- 5. Load Book3 template --------
    TEMPLATE = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "static", "templates", "Book3_template.xlsx")
    )
    try:
        template_df = pd.read_excel(TEMPLATE)
    except Exception:
        raise HTTPException(500, "Book3_template.xlsx missing on server")

    colA, colB, colC = template_df.columns

    wb = Workbook()
    ws = wb.active
    ws.title = "ATP11A"

    # first row: header row with column names
    ws.append([colA, colB, colC])

    ipv6_row_seen = False  # so only first IPv6 row gets comma-separated list
    azimuth_row_seen = False
    a6height_row_seen = False
    a6tilt_row_seen = False

    for _, row in template_df.iterrows():
        hc = "" if pd.isna(row[colA]) else str(row[colA]).strip()
        src = "" if pd.isna(row[colB]) else str(row[colB]).strip()
        new_val = row[colC]
        lower = hc.lower()

        # ----- dynamic replacements based on Column A text -----

        # PMP SAP ID
        if "pmp sap id" in lower:
            new_val = base_pmp
        elif "site/location name" in lower or "site/location address" in lower:
            new_val = base_sitename
        # A6 NE ID per sector
        elif "a6 ne id" in lower and "sect" in lower:
            sec_target = sector_from_hc(hc)
            new_val = a6_for_sector(sec_target)

        # MAC Address per sector
        elif "mac address of base terminal" in lower and "sect" in lower:
            sec_target = sector_from_hc(hc)
            new_val = next(
                (d["mac"] for d in sec_info if d["sector"] == sec_target),
                ""
            )

        # Serial Number per sector
        elif "serial number of base terminal" in lower and "sect" in lower:
            sec_target = sector_from_hc(hc)
            new_val = next(
                (d["rsn"] for d in sec_info if d["sector"] == sec_target),
                ""
            )

        # IPv6 pool – one combined cell
        elif "ipv6 pool address" in lower:
            sec_target = sector_from_hc(hc)
            new_val = a6ip_for_sector(sec_target)

                # --- Azimuth: one combined cell (comma-separated for all sectors) ---
        elif "base terminal actual azimuth (in degree)" in lower:
            if not azimuth_row_seen:
                new_val = azimuth_combined
                azimuth_row_seen = True
            else:
                new_val = ""  # keep other azimuth rows blank

        # --- Proposed A6 Height: one combined cell (repeat value N times) ---
        elif "base terminal actual height (in mtr)" in lower:
            if not a6height_row_seen:
                new_val = a6height_combined
                a6height_row_seen = True
            else:
                new_val = ""

        # --- Proposed A6 Tilt: one combined cell (repeat value N times) ---
        elif "base terminal actual tilt (in degree)" in lower:
            if not a6tilt_row_seen:
                new_val = a6tilt_combined
                a6tilt_row_seen = True
            else:
                new_val = ""


        # GIS Sector ID
        elif "gis sector" in lower:
            new_val = base_gis

        # eNB SAP ID (if you want to copy GIS or separate col, adjust here)
        elif "enb/css site sap id" in lower:
            sec_target = sector_from_hc(hc)

            if sec_target:
                # sect1 → suffix -1
                sec_num = int(re.findall(r"\d+", sec_target)[0])
                new_val = f"{base_gis}-{sec_num}"
            else:
                # main row without sector suffix
                new_val = base_gis


        # Circle
        elif lower.strip() == "circle":
            new_val = circle

        # everything else stays hard-coded from template

        ws.append([hc, src, new_val])

    # -------- 6. Formatting: yellow sections + bold headers --------
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    section_headers = {
        "Site Detail",
        "Installtion Details",   # as in template (typo)
        "Installation Details",  # safety
        "Cable",
        "Power Rating Parameter",
        "Radio Details",
        "Labelling",
        "Snap",
    }

    # Header row bold
    for col in ["A", "B", "C"]:
        ws[f"{col}1"].font = bold_font

    # Section header background & bold
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cellA = row[0]
        if cellA.value and str(cellA.value).strip() in section_headers:
            cellA.fill = yellow_fill
            cellA.font = bold_font

    # -------- 7. Return file --------
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"A6_HOTO_{base_sitename}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Filename": filename,  # ✅ easy for frontend to read
        "Access-Control-Expose-Headers": "Content-Disposition, X-Filename",  # ✅ IMPORTANT
    },
    )


# ------------------------------------------------------------
# XLSX (no images)
# ------------------------------------------------------------
@router.post("/jobs/{job_id}/export.xlsx")
async def export_xlsx(
    job_id: str,
    mainExcel: UploadFile = File(...),
    db=Depends(get_db)
):
    
    

    # -------- 1. Fetch job --------
    try:
        base_job = db.jobs.find_one({"_id": ObjectId(job_id)})
    except Exception:
        raise HTTPException(400, "Invalid Job ID")

    if not base_job:
        raise HTTPException(404, "Job not found")

    worker = base_job.get("workerPhone")
    site_id = str(base_job.get("siteId", "")).strip()
    circle = base_job.get("circle", "")

    # -------- 2. Collect all sectors (MAC/RSN) for this site --------
    related = list(db.jobs.find({"workerPhone": worker, "siteId": site_id}))
    sec_info = []

    for j in related:
        raw_sec = str(j.get("sector", "")).strip()

        # normalize sector → Sec1 / Sec2 / Sec3
        alpha_map = {"alpha": 1, "beta": 2, "gamma": 3}

        sec_num = None

        # case 1: Alpha/Beta/Gamma
        low = raw_sec.lower()
        if low in alpha_map:
            sec_num = alpha_map[low]

        # case 2: -1 / -2 / -3
        elif raw_sec.startswith("-") and raw_sec[1:].isdigit():
            sec_num = int(raw_sec[1:])

        # case 3: sector already numeric (1 / 2 / 3)
        elif raw_sec.isdigit():
            sec_num = int(raw_sec)

        # case 4: sector already in format Sec1 / Sec2
        elif raw_sec.lower().startswith("sec") and raw_sec[3:].isdigit():
            sec_num = int(raw_sec[3:])

        # create final normalized format
        sector_norm = f"Sec{sec_num}" if sec_num else raw_sec
        azimuth = j.get("azimuthDeg") or ""
        mac = j.get("macId") or ""
        rsn = j.get("rsnId") or ""

        # fallback from photos
        photos = db.photos.find({"jobId": str(j["_id"])})
        for p in photos:
            f = p.get("fields", {}) or {}
            if not mac and f.get("macId"):
                mac = f["macId"]
            if not rsn and f.get("rsn"):
                rsn = f["rsn"]
        sec_info.append({"sector": sector_norm, "mac": mac, "rsn": rsn, "azimuth": azimuth})


    # -------- 3. Read uploaded Main Excel --------
    try:
        df = pd.read_excel(mainExcel.file)
    except Exception:
        raise HTTPException(400, "Failed to read uploaded Excel")

    cols = {c.lower(): c for c in df.columns}

    def find_c(*keys):
        for low, orig in cols.items():
            if all(k.lower() in low for k in keys):
                return orig
        return None

    site_col = find_c("enbsiteid") or find_c("site")
    pmp_col = find_c("pmp sap id", "sap")
    a6_col = find_c("a6neid")
    gis_col = find_c("gis sector_id", "sector")
    a6ip_col = find_c("a6 ip")
    a6hieght_col = find_c("enb antenna height")
    a6tilt_col = find_c("proposed a6 tilt")
    sitename_col = find_c("site name")

    if not site_col:
        raise HTTPException(400, "Site / eNBsiteID not found in Main Excel")

    df_site = df[site_col].astype(str).str.strip()
    match = df[df_site == site_id]
    if match.empty:
        match = df[df_site.str.contains(site_id, case=False, na=False)]

    if match.empty:
        base_pmp = base_a6 = base_gis = base_a6ip = base_a6height = base_a6tilt = ""
    else:
        r = match.iloc[0]

        def safe(c):
            return "" if not c or c not in r or pd.isna(r[c]) else str(r[c])

        base_pmp = safe(pmp_col)
        base_a6 = safe(a6_col)
        base_gis = safe(gis_col)
        base_a6ip = safe(a6ip_col)
        base_a6height = safe(a6hieght_col)
        base_a6tilt = safe(a6tilt_col)
        base_sitename = safe(sitename_col)


    def _sec_sort_key(x):
        m = re.findall(r"\d+", str(x.get("sector", "")))
        return int(m[0]) if m else 999

    sec_info_sorted = sorted(sec_info, key=_sec_sort_key)
    sector_count = len(sec_info_sorted)

    # --- Azimuth: comma-separated across sectors (skip blanks) ---
    azimuth_values = [str(d.get("azimuth", "")).strip() for d in sec_info_sorted]
    azimuth_values = [v for v in azimuth_values if v]  # remove empty
    azimuth_combined = ", ".join(azimuth_values)

    # --- Height / Tilt: repeat same base value N times ---
    def repeat_base(val: str, n: int) -> str:
        v = (val or "").strip()
        if not v or n <= 0:
            return ""
        return ", ".join([v] * n)

    a6height_combined = repeat_base(base_a6height, sector_count)

    a6tilt_combined   = repeat_base(base_a6tilt, sector_count)
    def a6ip_for_sector(sec: str):
        if not base_a6ip:
            return ""

        # Extract last digits of base A6-IP
        m = re.search(r"(\d+)$", base_a6ip)
        if not m:
            return base_a6ip

        base_last = int(m.group(1))

        # Find all existing sector numbers
        sector_numbers = sorted(
            int(re.findall(r"\d+", d["sector"])[0])
            for d in sec_info
        )

        sec_num = int(re.findall(r"\d+", sec)[0])

        # If only one sector exists
        if len(sector_numbers) == 1:
            if sector_numbers[0] == sec_num:
                return base_a6ip
            return ""

        # Multi-sector case
        base_sec = sector_numbers[0]
        diff = sec_num - base_sec

        new_last = base_last + diff
        prefix = base_a6ip[:-len(str(base_last))]
        return prefix + str(new_last)


    # -------- 4. Function for A6 sequencing --------
    def a6_for_sector(sec: str):
        """
        Final A6 generation logic:
        - 1 sector: place base_a6 only for that sector
        - 2 sectors: match last digit, other gets prefix + sector number
        - 3 sectors: simple 1,2,3 mapping
        """
        if not base_a6 or not sec:
            return ""

        # extract target sector number
        m = re.findall(r"\d+", sec)
        if not m:
            return ""
        target_num = int(m[0])   # e.g. Sec3 → 3

        # extract all sectors from sec_info
        sector_nums = sorted(int(re.findall(r"\d+", d["sector"])[0]) for d in sec_info)
        count = len(sector_nums)

        # split prefix and numeric suffix
        t = re.search(r"(\d+)$", base_a6)
        if not t:
            return base_a6

        full_suffix = t.group(1)          # 6002
        suffix_digit = int(full_suffix[-1])  # 2
        prefix = base_a6[:-len(full_suffix)] # before "6002"

        # ----------------------------------------------------------------
        # CASE 1 — ONLY ONE SECTOR
        # ----------------------------------------------------------------
        if count == 1:
            only = sector_nums[0]
            if only == target_num:
                return base_a6
            return ""  # others empty

        # ----------------------------------------------------------------
        # CASE 2 — TWO SECTORS
        # ----------------------------------------------------------------
        if count == 2:
            s1, s2 = sector_nums   # two sector numbers

            # If the suffix digit matches one of the present sectors
            if suffix_digit in sector_nums:
                if target_num == suffix_digit:
                    return base_a6
                else:
                    # for the other sector → replace last digit
                    return f"{prefix}{full_suffix[:-1]}{target_num}"

            # suffix does NOT match → generate by replacing last digit
            return f"{prefix}{full_suffix[:-1]}{target_num}"

        # ----------------------------------------------------------------
        # CASE 3 — ALL THREE SECTORS
        # ----------------------------------------------------------------
        # Always: Sec1→6001, Sec2→6002, Sec3→6003
        return f"{prefix}{full_suffix[:-1]}{target_num}"



    # -------- 5. Load Template --------
    TEMPLATE = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "static", "templates", "Book1_template.xlsx")
    )
    try:
        template_df = pd.read_excel(TEMPLATE)
    except Exception:
        raise HTTPException(500, "Book1_template.xlsx missing on server")

    colA, colB, colC = template_df.columns  # we will ignore colD

    # -------- 6. Build Workbook (A, B, C only) --------
    wb = Workbook()
    ws = wb.active
    ws.title = "ATP11A"
    ws.append([colA, colB, colC])  # header without Business Rule
    azimuth_row_seen = False
    a6height_row_seen = False
    a6tilt_row_seen = False
    for _, row in template_df.iterrows():
        hc = "" if pd.isna(row[colA]) else str(row[colA]).strip()
        src = "" if pd.isna(row[colB]) else str(row[colB]).strip()
        if "hard coded structure" in hc.lower():
            continue

        # -------- Detect sector (NEW RULE) --------
        expected_sector = None
        
        if "sect" in src.lower():  # Sect1 / Sect2 / Sect3 inside Column B
            m = re.findall(r"\d+", src)
            expected_sector = None if not m else f"Sec{m[0]}"
        
        # -------- Get MAC / RSN --------
        mac = rsn = ""
        if expected_sector:
            for d in sec_info:
               
                if d["sector"] == expected_sector:
                    mac = d["mac"]
                    rsn = d["rsn"]
                    break


        # Default (preserve template hard-coded)
        new_val = row[colC]
        lower = hc.lower()

        # Replace only dynamic values
        # Replace only dynamic values
        if "pmp sap id" in lower:
            new_val = base_pmp
        elif "site/location name" in lower or "site/location address" in lower:
            new_val = base_sitename
        elif "a6 ne id" in lower:
            # A6 column logic considers sector from column B if available
           
            sec_target = expected_sector
            new_val = a6_for_sector(sec_target)

        elif "ipv6 pool address" in lower:

            # build comma-separated IP list for all sectors
            ip_list = []

            for d in sec_info:
                sec = d["sector"]
                ip_val = a6ip_for_sector(sec)
                if ip_val:
                    ip_list.append(ip_val)
            # put everything in ONE cell
            new_val = ", ".join(ip_list)



        elif "gis sector" in lower:
            new_val = base_gis

        elif "enb sap id" in lower or "enb/css site sap id" in lower:
            new_val = base_gis

        elif "base radio planned azimuth (in degree) (sect0,sect1,sect2)" in lower or "base radio actual azimuth (in degree) (sect0,sect1,sect2)" in lower:
            if not azimuth_row_seen:
                new_val = azimuth_combined
            else:
                new_val = ""  # keep other azimuth rows blank

        # --- Proposed A6 Height: one combined cell (repeat value N times) ---
        elif "base radio planned height (in mtr) (sect0,sect1,sect2)" in lower or "base radio actual height (in mtr) (sect0,sect1,sect2)" in lower:
            if not a6height_row_seen:
                new_val = a6height_combined
                print(" A6 Height combined:", new_val)
            else:
                new_val = ""

        # --- Proposed A6 Tilt: one combined cell (repeat value N times) ---
        elif "base radio actual tilt (in degree) (sect0,sect1,sect2)" in lower:
            if not a6tilt_row_seen:
                new_val = a6tilt_combined
                a6tilt_row_seen = True
            else:
                new_val = ""

        elif "mac address" in lower:
            sec_target = expected_sector
            mac = next((d["mac"] for d in sec_info if d["sector"] == sec_target), "")
            new_val = mac
            
        elif "serial number" in lower:
            sec_target = expected_sector
            rsn = next((d["rsn"] for d in sec_info if d["sector"] == sec_target), "")
            new_val = rsn
            

        elif "circle" in lower:
            new_val = circle


        ws.append([hc, src, new_val])
    

# Yellow fill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    # Headers to highlight in Column A
    section_headers = {
        "Site Detail",
        "Installtion Details",
        "Cable",
        "Base Radio Details",
        "Labelling",
        "Snap"
    }

    # Make Row 1 bold (Columns A, B, C)
    for col in ["A", "B", "C"]:
        ws[f"{col}1"].font = bold_font

    # Apply formatting to matching rows
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cellA = row[0]   # Column A cell

        if cellA.value and str(cellA.value).strip() in section_headers:
            cellA.fill = yellow_fill
            cellA.font = bold_font

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"A6_{base_sitename}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Filename": filename,  # ✅ easy for frontend to read
        "Access-Control-Expose-Headers": "Content-Disposition, X-Filename",  # ✅ IMPORTANT
    },
    )







# ------------------------------------------------------------
# JOB ZIP (images)
# ------------------------------------------------------------
@router.get("/jobs/{job_id}/export.zip")
def export_job_zip(job_id: str,
                   db=Depends(get_db)):
    try:
        _id = ObjectId(job_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid job id")

    job = db.jobs.find_one({"_id": _id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    job_sector = job.get("sector") 

    photo_q = {"jobId": {"$in": [job_id, str(_id)]}}
    photos = list(db.photos.find(photo_q).sort("_id", 1))
    if not photos:
        raise HTTPException(status_code=404, detail="No photos for this job")
    
    def _clean_key(k: str | None) -> str | None:
        if not k:
            return None
        k = str(k)
        if k.startswith("s3://"):
            try:
                return k.split("/", 3)[-1].split("/", 1)[-1]
            except Exception:
                return None
        return k

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in photos:
            folder = f"Sec{job_sector}" if job_sector else "Unknown"

            base = (p.get("type") or "PHOTO").lower()
            key_raw = p.get("s3Key") or ""
            key = _clean_key(key_raw)

            ext = ".jpg"
            if key:
                low = key.lower()
                for e in (".jpeg", ".jpg", ".png", ".webp"):
                    if low.endswith(e):
                        ext = e
                        break
            
            p_sector = p.get("sector") or job_sector
            logical = f"sec{p_sector}_{base}{ext}"
            arcname = f"{folder}/{logical}"

            lp = p.get("localPath")
            if lp and os.path.exists(lp):
                zf.write(lp, arcname=arcname)
                continue
            if key:
                try:
                    url = presign_url(key, expires=3600)
                    with httpx.Client(timeout=20) as client:
                        r = client.get(url)
                        r.raise_for_status()
                        zf.writestr(arcname, r.content)
                        continue
                except Exception as ex:
                    print(f"[ZIP] fetch failed for key={key}: {ex}")
            zf.writestr(arcname.replace(ext, "_MISSING.txt"), b"Missing or inaccessible image")


    mem.seek(0)
    fname = f'job_{job_id}_sec{job_sector}.zip'
    return StreamingResponse(
        mem,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )



# ------------------------------------------------------------
# TEMPLATE FOR SECTOR (single)
# ------------------------------------------------------------
@router.get("/jobs/templates/sector/{sector}")
def job_template(sector: str):
    types = build_required_types_for_sector(sector)
    return {
        "requiredTypes": types,
        "labels": {t: type_label(t) for t in types},
        "sector": sector,
    }


# ------------------------------------------------------------
# MANUAL EXPORT – SECTOR-WISE EXCEL (one sheet per sector)
# ------------------------------------------------------------
def _dt_or_none(s: str | None):
    if not s:
        return None
    return dt.datetime.strptime(s, "%Y-%m-%d")


@router.get("/exports/sector.xlsx")
def export_sector_xlsx(
    db=Depends(get_db),
    date_from: str | None = Query(None, description="YYYY-MM-DD"),
    date_to: str | None = Query(None, description="YYYY-MM-DD"),
):
    """
    Manual export – data grouped by sector, one sheet per sector.
    """
    q = {}
    if date_from or date_to:
        tmin = _dt_or_none(date_from) or dt.datetime.min
        tmax = _dt_or_none(date_to) or dt.datetime.max
        q["createdAt"] = {"$gte": tmin, "$lte": tmax}

    jobs = list(db.jobs.find(q))
    by_sector: Dict[str | None, list[dict]] = {}

    for j in jobs:
        job_id = str(j["_id"])
        worker = j.get("workerPhone")
        photos = list(db.photos.find({"jobId": job_id}))
        for p in photos:
            photo_sector = str(p.get("sector")) if p.get("sector") is not None else None
            f = p.get("fields", {}) or {}
            c = p.get("checks", {}) or {}
            base = (p.get("type") or "PHOTO").lower()
            
            ext = ".jpg"
            key = p.get("s3Key", "")
            for e in (".jpeg", ".jpg", ".png", ".webp"):
                if key.lower().endswith(e):
                    ext = e
                    break
            logical = (f"sec{photo_sector}_{base}{ext}"
                       if photo_sector is not None else f"{base}{ext}")

            row = {
                "jobId": job_id,
                "workerPhone": worker,
                "sector": photo_sector,
                "photoId": str(p.get("_id")),
                "type": p.get("type"),
                "s3Key": key,
                "s3Url": presign_url(key) if key else None,
                "logicalName": logical,
                "macId": f.get("macId"),
                "rsn": f.get("rsn"),
                "azimuthDeg": f.get("azimuthDeg"),
                "blurScore": c.get("blurScore"),
                "isDuplicate": c.get("isDuplicate"),
                "skewDeg": c.get("skewDeg"),
                "status": p.get("status"),
                "reason": "|".join(p.get("reason") or []),
            }
            by_sector.setdefault(photo_sector, []).append(row)

    # Build workbook (one sheet per sector)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def write_sheet(title: str, rows: list[dict]):
        ws = wb.create_sheet(title=title)
        if not rows:
            ws.append(["No data"])
            return
        df = pd.DataFrame(rows)
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append(list(r.values))

    for s in sorted(k for k in by_sector.keys() if k is not None):
        write_sheet(f"Sec{s}", by_sector[s])
    if None in by_sector:
        write_sheet("Unknown", by_sector[None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        headers={"Content-Disposition": 'attachment; filename="export_sector.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )